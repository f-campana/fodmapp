#!/usr/bin/env python3
"""
CIQUAL 2025 -> FODMAP Planner ETL
=================================
Hybrid ingestion for FR-first loading:
- XLSX provides nutrient values (fructose, galactose, glucose, lactose, total polyols, sugars)
- XML (alim + alim_grp) provides canonical FR names and category hierarchy

Usage:
  python ciqual_etl.py inspect path/to/Table_Ciqual_2025_ENG.xlsx
  python ciqual_etl.py stats path/to/Table_Ciqual_2025_ENG.xlsx
  python ciqual_etl.py load path/to/Table_Ciqual_2025_ENG.xlsx \
    --alim-xml path/to/alim_2025_11_03.xml \
    --alim-grp-xml path/to/alim_grp_2025_11_03.xml \
    --db-url postgresql://user:pass@localhost:5432/fodmap_test

Requirements:
  pip install openpyxl psycopg2-binary
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import unicodedata
import xml.etree.ElementTree as ET
from datetime import date
from pathlib import Path

import openpyxl


TARGET_CONSTITUENTS = {
    "32210": {"nutrient_code": "CIQUAL_32210", "name": "Fructose", "infoods": "FRUS"},
    "32220": {"nutrient_code": "CIQUAL_32220", "name": "Galactose", "infoods": "GALS"},
    "32250": {"nutrient_code": "CIQUAL_32250", "name": "Glucose", "infoods": "GLUS"},
    "32410": {"nutrient_code": "CIQUAL_32410", "name": "Lactose", "infoods": "LACS"},
    "34000": {"nutrient_code": "CIQUAL_34000", "name": "Total polyols", "infoods": "POLYL"},
    "32000": {"nutrient_code": "CIQUAL_32000", "name": "Sugars", "infoods": "SUGAR"},
}

HEADER_PATTERNS = {
    "32210": [r"\b32210\b", r"\bfructose\b"],
    "32220": [r"\b32220\b", r"\bgalactose\b"],
    "32250": [r"\b32250\b", r"\bglucose\b"],
    "32410": [r"\b32410\b", r"\blactose\b"],
    "34000": [r"\b34000\b", r"\bpolyols\b"],
    "32000": [r"\b32000\b", r"\bsugars?\b"],
}

FOOD_ID_PATTERNS = {
    "alim_code": [r"\balim[_\s]*code\b", r"\bfood[_\s]*code\b", r"^code$"],
    "alim_nom_fr": [r"\balim[_\s]*nom[_\s]*fr\b", r"\bnom[_\s]*fran[çc]ais\b", r"\bfrench[_\s]*name\b"],
    "alim_nom_eng": [r"\balim[_\s]*nom[_\s]*eng\b", r"\benglish[_\s]*name\b", r"\bfood[_\s]*name\b", r"\bname\b"],
    "alim_nom_sci": [r"\balim[_\s]*nom[_\s]*sci\b", r"\bscientific[_\s]*name\b"],
    "alim_grp_code": [r"\balim[_\s]*grp[_\s]*code\b", r"\bgroup[_\s]*code\b"],
    "alim_grp_nom_fr": [r"\balim[_\s]*grp[_\s]*nom[_\s]*fr\b", r"\bgroup[_\s]*name[_\s]*fr\b"],
    "alim_grp_nom_eng": [r"\balim[_\s]*grp[_\s]*nom[_\s]*eng\b", r"\bgroup[_\s]*name\b", r"\bfood[_\s]*group\b"],
    "alim_ssgrp_code": [r"\balim[_\s]*ssgrp[_\s]*code\b", r"\bsubgroup[_\s]*code\b"],
    "alim_ssgrp_nom_fr": [r"\balim[_\s]*ssgrp[_\s]*nom[_\s]*fr\b", r"\bsubgroup[_\s]*name[_\s]*fr\b"],
    "alim_ssgrp_nom_eng": [r"\balim[_\s]*ssgrp[_\s]*nom[_\s]*eng\b", r"\bsubgroup[_\s]*name\b", r"\bfood[_\s]*sub\s*group\b"],
    "alim_ssssgrp_code": [r"\balim[_\s]*ssssgrp[_\s]*code\b", r"\bsubsubgroup[_\s]*code\b"],
    "alim_ssssgrp_nom_fr": [r"\balim[_\s]*ssssgrp[_\s]*nom[_\s]*fr\b", r"\bsubsubgroup[_\s]*name[_\s]*fr\b"],
    "alim_ssssgrp_nom_eng": [r"\balim[_\s]*ssssgrp[_\s]*nom[_\s]*eng\b", r"\bsubsubgroup[_\s]*name\b"],
}

IDENTITY_FIELDS = [
    "alim_nom_fr",
    "alim_nom_eng",
    "alim_nom_sci",
    "alim_grp_code",
    "alim_grp_nom_fr",
    "alim_grp_nom_eng",
    "alim_ssgrp_code",
    "alim_ssgrp_nom_fr",
    "alim_ssgrp_nom_eng",
    "alim_ssssgrp_code",
    "alim_ssssgrp_nom_fr",
    "alim_ssssgrp_nom_eng",
]

HEADER_SCAN_ROWS = 12

DEFAULT_DB_URL = f"postgresql://{os.getenv('USER', 'postgres')}@localhost:5432/fodmap_test"
CIQUAL_SOURCE_SLUG = "ciqual_2025"

CIQUAL_OBSERVED_AT = date(2025, 11, 3)
CIQUAL_EFFECTIVE_FROM = date(2025, 11, 19)

PLACEHOLDER_FR_RE = re.compile(r"^CIQUAL\s+\d+$", re.IGNORECASE)


class LoadResult:
    def __init__(self):
        self.foods_created = 0
        self.foods_existing = 0
        self.foods_unresolved_fr = 0
        self.observations_inserted = 0
        self.observations_skipped = 0
        self.categories_assigned = 0
        self.aggregate_rows_skipped = 0
        self.categories_created = 0
        self.categories_updated = 0
        self.errors = []
        self.unresolved_examples = []
        self.unresolved_count = 0
        self.exit_code = 0


def _normalize_header(value):
    text = "" if value is None else str(value)
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def _clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text or text == "-":
        return None
    return text


def _normalize_code(value, width=None):
    text = _clean_text(value)
    if text is None:
        return None
    if width is not None and text.isdigit():
        text = text.zfill(width)
    return text


def _get_cell(row_values, col_idx):
    if col_idx is None:
        return None
    pos = col_idx - 1
    if pos < 0 or pos >= len(row_values):
        return None
    return row_values[pos]


def _clean_name(value):
    return _clean_text(value)


def _make_slug(name):
    s = unicodedata.normalize("NFKD", name)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = s.strip("-")
    s = re.sub(r"-+", "-", s)
    return s[:100]


def _is_placeholder_ciqual_name(name):
    if name is None:
        return False
    return bool(PLACEHOLDER_FR_RE.match(name.strip()))


def _print_mapping(metadata):
    col_map = metadata["col_map"]
    nutrients_found = metadata["nutrients_found"]
    print(f"  Header row: {metadata['header_row']}")
    print(f"  Food code column: {col_map.get('alim_code')}")
    print(f"  Food name FR column: {col_map.get('alim_nom_fr', 'NOT FOUND')}")
    print(f"  Food name EN column: {col_map.get('alim_nom_eng', 'NOT FOUND')}")
    print(f"  Target nutrients found: {len(nutrients_found)}/{len(TARGET_CONSTITUENTS)}")
    for const_code, col_idx in nutrients_found.items():
        info = TARGET_CONSTITUENTS[const_code]
        label = metadata["header_texts"].get(col_idx, "")
        print(f"    {info['name']} ({const_code}) -> column {col_idx}: \"{label}\"")

    missing = set(TARGET_CONSTITUENTS.keys()) - set(nutrients_found.keys())
    if missing:
        print(f"  WARNING missing nutrients: {', '.join(sorted(missing))}")


def parse_ciqual_value(raw):
    """
    Parse a CIQUAL cell value into (comparator, amount_value, amount_raw).

    CIQUAL conventions:
    - French decimal commas (0,3)
    - comparators (< 0,2)
    - markers (traces, nd, -)
    """
    if raw is None:
        return ("missing", None, "")

    if isinstance(raw, (int, float)):
        return ("eq", round(float(raw), 6), str(raw))

    raw_str = str(raw).strip()
    if raw_str == "" or raw_str == "-":
        return ("missing", None, raw_str if raw_str else "-")

    raw_lower = raw_str.lower()
    if raw_lower in ("traces", "trace", "tr"):
        return ("trace", None, raw_str)
    if raw_lower in ("nd", "n.d.", "n.d"):
        return ("nd", None, raw_str)

    match = re.match(r"^<=\s*(.+)$", raw_str)
    if match:
        val = _parse_number(match.group(1))
        return ("lte", val, raw_str) if val is not None else ("missing", None, raw_str)

    match = re.match(r"^>=\s*(.+)$", raw_str)
    if match:
        val = _parse_number(match.group(1))
        return ("gte", val, raw_str) if val is not None else ("missing", None, raw_str)

    match = re.match(r"^<\s*(.+)$", raw_str)
    if match:
        val = _parse_number(match.group(1))
        return ("lt", val, raw_str) if val is not None else ("missing", None, raw_str)

    match = re.match(r"^>\s*(.+)$", raw_str)
    if match:
        val = _parse_number(match.group(1))
        return ("gt", val, raw_str) if val is not None else ("missing", None, raw_str)

    val = _parse_number(raw_str)
    if val is not None:
        return ("eq", val, raw_str)

    return ("missing", None, raw_str)


def _parse_number(text):
    text = text.strip().replace(",", ".").replace(" ", "")
    try:
        return round(float(text), 6)
    except (TypeError, ValueError):
        return None


def find_header_row(rows, max_columns):
    """
    Find the best header row from pre-read rows.
    Returns (header_row_index, header_map).
    header_map uses 1-based column indices.
    """
    best_row = None
    best_count = 0
    best_headers = {}

    for row_idx, row_values in enumerate(rows, start=1):
        headers = {}
        count = 0
        for col_idx in range(1, max_columns + 1):
            val = row_values[col_idx - 1] if col_idx - 1 < len(row_values) else None
            if val is None:
                continue
            text = str(val).strip()
            if not text:
                continue
            headers[col_idx] = text
            count += 1

        has_food_code = any(
            any(re.search(pattern, _normalize_header(h), re.IGNORECASE) for pattern in FOOD_ID_PATTERNS["alim_code"])
            for h in headers.values()
        )

        if has_food_code and count > best_count:
            best_row = row_idx
            best_count = count
            best_headers = headers

    return best_row, best_headers


def map_columns(headers):
    """Map header text to known fields and nutrient columns."""
    mapping = {}

    for col_idx, header_text in headers.items():
        header_norm = _normalize_header(header_text)

        for field_name, patterns in FOOD_ID_PATTERNS.items():
            if field_name in mapping:
                continue
            if any(re.search(pattern, header_norm, re.IGNORECASE) for pattern in patterns):
                mapping[field_name] = col_idx
                break

        for const_code, patterns in HEADER_PATTERNS.items():
            if const_code in mapping:
                continue
            if any(re.search(pattern, header_norm, re.IGNORECASE) for pattern in patterns):
                mapping[const_code] = col_idx
                break

    return mapping


def _select_sheet(wb):
    if "food composition" in wb.sheetnames:
        return wb["food composition"]
    if wb.active is not None:
        return wb.active
    return wb.worksheets[0]


def open_ciqual_workbook(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = _select_sheet(wb)

    max_scan = min(HEADER_SCAN_ROWS, ws.max_row)
    scan_rows = list(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True))
    header_row, headers = find_header_row(scan_rows, ws.max_column)
    if header_row is None:
        wb.close()
        raise ValueError("Could not find header row in the workbook")

    col_map = map_columns(headers)
    if "alim_code" not in col_map:
        wb.close()
        raise ValueError("Could not find alim_code column")

    nutrients_found = {k: v for k, v in col_map.items() if k in TARGET_CONSTITUENTS}
    metadata = {
        "sheet": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "header_row": header_row,
        "header_texts": headers,
        "col_map": col_map,
        "nutrients_found": nutrients_found,
    }
    return wb, ws, metadata


def iter_ciqual_records(ws, metadata):
    """Yield parsed food records from CIQUAL workbook rows."""
    col_map = metadata["col_map"]
    nutrients_found = metadata["nutrients_found"]
    header_row = metadata["header_row"]

    for row_values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        alim_code = _normalize_code(_get_cell(row_values, col_map["alim_code"]))
        if not alim_code:
            continue

        record = {
            "alim_code": alim_code,
            "nutrients": {},
        }
        for field in IDENTITY_FIELDS:
            record[field] = _clean_text(_get_cell(row_values, col_map.get(field)))

        for const_code, col_idx in nutrients_found.items():
            raw_val = _get_cell(row_values, col_idx)
            comparator, amount, raw_str = parse_ciqual_value(raw_val)
            record["nutrients"][const_code] = {
                "comparator": comparator,
                "amount_value": amount,
                "amount_raw": raw_str,
                "nutrient_code": TARGET_CONSTITUENTS[const_code]["nutrient_code"],
            }

        yield record


def load_alim_xml(filepath):
    """Load per-food FR/EN identity + hierarchy from CIQUAL alim XML."""
    root = ET.parse(filepath).getroot()
    items = {}

    for row in root.findall(".//ALIM"):
        alim_code = _normalize_code(row.findtext("alim_code"))
        if not alim_code:
            continue

        items[alim_code] = {
            "alim_nom_fr": _clean_name(row.findtext("alim_nom_fr")),
            "alim_nom_eng": _clean_name(row.findtext("alim_nom_eng")),
            "alim_nom_sci": _clean_name(row.findtext("alim_nom_sci")),
            "alim_grp_code": _normalize_code(row.findtext("alim_grp_code"), width=2),
            "alim_ssgrp_code": _normalize_code(row.findtext("alim_ssgrp_code"), width=4),
            "alim_ssssgrp_code": _normalize_code(row.findtext("alim_ssssgrp_code"), width=6),
        }

    return items


def load_alim_grp_xml(filepath):
    """
    Load CIQUAL category hierarchy definitions from alim_grp XML.
    Returns node dict keyed by deterministic SQL code.
    """
    root = ET.parse(filepath).getroot()
    nodes = {}

    for row in root.findall(".//ALIM_GRP"):
        grp_code = _normalize_code(row.findtext("alim_grp_code"), width=2)
        grp_fr = _clean_name(row.findtext("alim_grp_nom_fr"))
        grp_en = _clean_name(row.findtext("alim_grp_nom_eng"))
        ssgrp_code = _normalize_code(row.findtext("alim_ssgrp_code"), width=4)
        ssgrp_fr = _clean_name(row.findtext("alim_ssgrp_nom_fr"))
        ssgrp_en = _clean_name(row.findtext("alim_ssgrp_nom_eng"))
        ssss_code = _normalize_code(row.findtext("alim_ssssgrp_code"), width=6)
        ssss_fr = _clean_name(row.findtext("alim_ssssgrp_nom_fr"))
        ssss_en = _clean_name(row.findtext("alim_ssssgrp_nom_eng"))

        if grp_code:
            grp_key = f"ciqual_grp_{grp_code}"
            nodes[grp_key] = {
                "code": grp_key,
                "parent_code": None,
                "level": 1,
                "name_fr": grp_fr,
                "name_en": grp_en,
            }

            if ssgrp_code and ssgrp_code != "0000":
                ssgrp_key = f"ciqual_ssgrp_{ssgrp_code}"
                nodes[ssgrp_key] = {
                    "code": ssgrp_key,
                    "parent_code": grp_key,
                    "level": 2,
                    "name_fr": ssgrp_fr,
                    "name_en": ssgrp_en,
                }

                if ssss_code and ssss_code != "000000":
                    ssss_key = f"ciqual_ssssgrp_{ssss_code}"
                    nodes[ssss_key] = {
                        "code": ssss_key,
                        "parent_code": ssgrp_key,
                        "level": 3,
                        "name_fr": ssss_fr,
                        "name_en": ssss_en,
                    }

    return nodes


def merge_record_identity(record, alim_lookup):
    """XML-first merge for identity and hierarchy fields."""
    xml_rec = alim_lookup.get(record["alim_code"], {})
    merged = {
        "alim_code": record["alim_code"],
        "nutrients": record["nutrients"],
    }

    for field in IDENTITY_FIELDS:
        merged[field] = xml_rec.get(field) or record.get(field)

    merged["alim_grp_code"] = _normalize_code(merged.get("alim_grp_code"), width=2)
    merged["alim_ssgrp_code"] = _normalize_code(merged.get("alim_ssgrp_code"), width=4)
    merged["alim_ssssgrp_code"] = _normalize_code(merged.get("alim_ssssgrp_code"), width=6)
    return merged


def _fallback_category_name(level, code, name_fr, name_en):
    fr = name_fr
    en = name_en
    if not fr:
        if level == 1:
            fr = f"Groupe CIQUAL {code}"
        elif level == 2:
            fr = f"Sous-groupe CIQUAL {code}"
        else:
            fr = f"Sous-sous-groupe CIQUAL {code}"
    if not en:
        if level == 1:
            en = f"CIQUAL group {code}"
        elif level == 2:
            en = f"CIQUAL subgroup {code}"
        else:
            en = f"CIQUAL subsubgroup {code}"
    return fr, en


def _upsert_category(cur, code, parent_category_id, level, name_fr, name_en):
    cur.execute("SELECT category_id FROM food_categories WHERE code = %s", (code,))
    row = cur.fetchone()
    if row:
        category_id = row[0]
        cur.execute(
            """
            UPDATE food_categories
            SET parent_category_id = %s,
                name_fr = %s,
                name_en = %s,
                level = %s,
                source_system = 'ciqual_2025'
            WHERE category_id = %s
            """,
            (parent_category_id, name_fr, name_en, level, category_id),
        )
        return category_id, False

    cur.execute(
        """
        INSERT INTO food_categories (code, parent_category_id, name_fr, name_en, level, source_system)
        VALUES (%s, %s, %s, %s, %s, 'ciqual_2025')
        RETURNING category_id
        """,
        (code, parent_category_id, name_fr, name_en, level),
    )
    return cur.fetchone()[0], True


def seed_ciqual_categories(cur, grp_nodes):
    """Upsert CIQUAL taxonomy nodes from alim_grp XML and return category_id by code."""
    by_code = {}
    created = 0
    updated = 0

    ordered = sorted(grp_nodes.values(), key=lambda n: (n["level"], n["code"]))
    for node in ordered:
        parent_id = by_code.get(node["parent_code"])
        level = node["level"]
        node_code = node["code"]

        suffix = node_code.rsplit("_", 1)[-1]
        name_fr, name_en = _fallback_category_name(level, suffix, node.get("name_fr"), node.get("name_en"))

        category_id, was_created = _upsert_category(cur, node_code, parent_id, level, name_fr, name_en)
        by_code[node_code] = category_id
        if was_created:
            created += 1
        else:
            updated += 1

    return by_code, created, updated


def ensure_category_chain_for_record(cur, category_ids, record):
    """Ensure a category chain exists for this food; return deepest category_id or None."""
    grp_code = record.get("alim_grp_code")
    ssgrp_code = record.get("alim_ssgrp_code")
    ssss_code = record.get("alim_ssssgrp_code")

    if not grp_code:
        return None

    grp_key = f"ciqual_grp_{grp_code}"
    ssgrp_key = f"ciqual_ssgrp_{ssgrp_code}" if ssgrp_code and ssgrp_code != "0000" else None
    ssss_key = f"ciqual_ssssgrp_{ssss_code}" if ssss_code and ssss_code != "000000" else None

    if grp_key not in category_ids:
        fr, en = _fallback_category_name(1, grp_code, record.get("alim_grp_nom_fr"), record.get("alim_grp_nom_eng"))
        category_id, _ = _upsert_category(cur, grp_key, None, 1, fr, en)
        category_ids[grp_key] = category_id

    if ssgrp_key and ssgrp_key not in category_ids:
        fr, en = _fallback_category_name(2, ssgrp_code, record.get("alim_ssgrp_nom_fr"), record.get("alim_ssgrp_nom_eng"))
        category_id, _ = _upsert_category(cur, ssgrp_key, category_ids[grp_key], 2, fr, en)
        category_ids[ssgrp_key] = category_id

    if ssss_key and ssss_key not in category_ids:
        fr, en = _fallback_category_name(3, ssss_code, record.get("alim_ssssgrp_nom_fr"), record.get("alim_ssssgrp_nom_eng"))
        parent_id = category_ids.get(ssgrp_key) if ssgrp_key else category_ids[grp_key]
        category_id, _ = _upsert_category(cur, ssss_key, parent_id, 3, fr, en)
        category_ids[ssss_key] = category_id

    if ssss_key and ssss_key in category_ids:
        return category_ids[ssss_key]
    if ssgrp_key and ssgrp_key in category_ids:
        return category_ids[ssgrp_key]
    return category_ids[grp_key]


def _ensure_unique_slug(cur, seed):
    slug = _make_slug(seed)
    if not slug:
        slug = "ciqual-food"

    candidate = slug
    suffix = 0
    while True:
        cur.execute("SELECT 1 FROM foods WHERE food_slug = %s", (candidate,))
        if cur.fetchone() is None:
            return candidate
        suffix += 1
        candidate = f"{slug}_{suffix}"


def _ensure_food_names(cur, food_id, source_id, name_fr, name_en):
    if name_fr:
        cur.execute(
            """
            INSERT INTO food_names (food_id, locale_code, name, is_primary, source_id)
            VALUES (%s, 'fr-FR', %s, TRUE, %s)
            ON CONFLICT DO NOTHING
            """,
            (food_id, name_fr, source_id),
        )
    if name_en:
        cur.execute(
            """
            INSERT INTO food_names (food_id, locale_code, name, is_primary, source_id)
            VALUES (%s, 'en-GB', %s, TRUE, %s)
            ON CONFLICT DO NOTHING
            """,
            (food_id, name_en, source_id),
        )


def cmd_inspect(filepath):
    print("=== CIQUAL 2025 File Inspection ===")
    print(f"File: {filepath}")
    print()

    wb, ws, metadata = open_ciqual_workbook(filepath)
    try:
        print(f"Sheet: {metadata['sheet']}")
        print(f"Dimensions: {metadata['max_row']} rows x {metadata['max_column']} columns")
        print(f"Header row: {metadata['header_row']}")
        print(f"Total mapped columns: {len(metadata['col_map'])}")
        print()

        print("=== Mapped columns ===")
        for key, col_idx in sorted(metadata["col_map"].items(), key=lambda x: x[1]):
            label = TARGET_CONSTITUENTS.get(key, {}).get("name", key)
            raw_header = metadata["header_texts"].get(col_idx, "")
            print(f"  Column {col_idx:3d}: {label:20s} <- \"{raw_header}\"")

        print()
        print("=== Sample data (first 5 food rows) ===")
        shown = 0
        for rec in iter_ciqual_records(ws, metadata):
            label = rec.get("alim_nom_eng") or rec.get("alim_nom_fr") or "?"
            print(f"  [{rec['alim_code']}] {label}")
            for const_code in TARGET_CONSTITUENTS:
                if const_code not in rec["nutrients"]:
                    continue
                obs = rec["nutrients"][const_code]
                info = TARGET_CONSTITUENTS[const_code]
                print(f"    {info['name']:15s}: raw={repr(obs['amount_raw']):15s} -> {obs['comparator']}({obs['amount_value']})")
            shown += 1
            if shown >= 5:
                break
    finally:
        wb.close()


def cmd_stats(filepath):
    print("=== CIQUAL 2025 Nutrient Coverage Stats ===")
    print(f"File: {filepath}")
    print()

    wb, ws, metadata = open_ciqual_workbook(filepath)
    try:
        counters = {
            code: {"eq": 0, "lt": 0, "lte": 0, "trace": 0, "missing": 0, "nd": 0, "other": 0}
            for code in TARGET_CONSTITUENTS
        }

        total = 0
        for record in iter_ciqual_records(ws, metadata):
            total += 1
            for const_code, obs in record["nutrients"].items():
                comp = obs["comparator"]
                if comp in counters[const_code]:
                    counters[const_code][comp] += 1
                else:
                    counters[const_code]["other"] += 1

        print(f"Total foods: {total}")
        print()
        print(f"{'Nutrient':<20s} {'numeric':>8s} {'<value':>8s} {'trace':>8s} {'missing':>8s} {'nd':>8s} {'coverage':>10s}")
        print("-" * 80)
        for const_code, info in TARGET_CONSTITUENTS.items():
            c = counters.get(const_code, {})
            numeric = c.get("eq", 0) + c.get("lt", 0) + c.get("lte", 0)
            trace = c.get("trace", 0)
            missing = c.get("missing", 0)
            nd = c.get("nd", 0)
            pct = (numeric + trace) / total * 100 if total > 0 else 0
            print(f"{info['name']:<20s} {numeric:>8d} {c.get('lt', 0):>8d} {trace:>8d} {missing:>8d} {nd:>8d} {pct:>9.1f}%")
    finally:
        wb.close()


def _resolve_source_and_nutrients(cur):
    cur.execute("SELECT source_id FROM sources WHERE source_slug = %s", (CIQUAL_SOURCE_SLUG,))
    row = cur.fetchone()
    if row is None:
        raise ValueError(f"Source '{CIQUAL_SOURCE_SLUG}' not found in sources table. Run schema DDL first.")
    source_id = row[0]

    nutrient_ids = {}
    for const_code, info in TARGET_CONSTITUENTS.items():
        cur.execute("SELECT nutrient_id FROM nutrient_definitions WHERE nutrient_code = %s", (info["nutrient_code"],))
        row = cur.fetchone()
        if row is None:
            print(f"  WARNING nutrient {info['nutrient_code']} not found in DB; skipping")
            continue
        nutrient_ids[const_code] = row[0]
        print(f"  Nutrient {info['name']:15s} ({info['nutrient_code']}) -> ID {row[0]}")

    return source_id, nutrient_ids


def _upsert_food(cur, source_id, alim_code, name_fr, name_en, scientific_name):
    cur.execute(
        """
        SELECT food_id
        FROM food_external_refs
        WHERE ref_system = 'CIQUAL' AND ref_value = %s
        ORDER BY food_id
        LIMIT 1
        """,
        (alim_code,),
    )
    row = cur.fetchone()

    if row:
        food_id = row[0]
        cur.execute(
            "SELECT canonical_name_fr, canonical_name_en, scientific_name FROM foods WHERE food_id = %s",
            (food_id,),
        )
        current_fr, current_en, current_sci = cur.fetchone()

        next_fr = current_fr
        next_en = current_en
        next_sci = current_sci

        if name_fr and (not current_fr or _is_placeholder_ciqual_name(current_fr)):
            next_fr = name_fr
        elif current_fr is None and name_fr is None:
            next_fr = None

        if name_en and not current_en:
            next_en = name_en
        if scientific_name and not current_sci:
            next_sci = scientific_name

        if (next_fr, next_en, next_sci) != (current_fr, current_en, current_sci):
            cur.execute(
                """
                UPDATE foods
                SET canonical_name_fr = %s,
                    canonical_name_en = %s,
                    scientific_name = %s,
                    updated_at = now()
                WHERE food_id = %s
                """,
                (next_fr, next_en, next_sci, food_id),
            )
        return food_id, False

    slug_seed = name_fr or name_en or f"ciqual-{alim_code}"
    food_slug = _ensure_unique_slug(cur, slug_seed)

    cur.execute(
        """
        INSERT INTO foods (food_slug, canonical_name_fr, canonical_name_en, scientific_name, preparation_state, status)
        VALUES (%s, %s, %s, %s, 'unknown', 'draft')
        RETURNING food_id
        """,
        (food_slug, name_fr, name_en, scientific_name),
    )
    food_id = cur.fetchone()[0]

    cur.execute(
        """
        INSERT INTO food_external_refs (food_id, ref_system, ref_value, source_id, country_code)
        VALUES (%s, 'CIQUAL', %s, %s, 'FR')
        ON CONFLICT DO NOTHING
        """,
        (food_id, alim_code, source_id),
    )

    return food_id, True


def _load_record(cur, source_id, nutrient_ids, category_ids, record, result):
    alim_code = record["alim_code"]
    name_fr = _clean_name(record.get("alim_nom_fr"))
    name_en = _clean_name(record.get("alim_nom_eng"))
    scientific_name = _clean_name(record.get("alim_nom_sci"))

    if name_fr is None:
        result.foods_unresolved_fr += 1
        if len(result.unresolved_examples) < 20:
            result.unresolved_examples.append((alim_code, name_en or ""))

    food_id = None

    cur.execute("SAVEPOINT sp_food")
    try:
        food_id, created = _upsert_food(cur, source_id, alim_code, name_fr, name_en, scientific_name)
        _ensure_food_names(cur, food_id, source_id, name_fr, name_en)
        cur.execute("RELEASE SAVEPOINT sp_food")
        if created:
            result.foods_created += 1
        else:
            result.foods_existing += 1
    except Exception as exc:
        cur.execute("ROLLBACK TO SAVEPOINT sp_food")
        cur.execute("RELEASE SAVEPOINT sp_food")
        result.errors.append(f"Food upsert {alim_code}: {exc}")
        return

    grp_code = record.get("alim_grp_code")
    if grp_code == "00":
        result.aggregate_rows_skipped += 1
    else:
        cur.execute("SAVEPOINT sp_category")
        try:
            target_category_id = ensure_category_chain_for_record(cur, category_ids, record)
            if target_category_id is not None:
                cur.execute(
                    """
                    DELETE FROM food_category_memberships fcm
                    USING food_categories fc
                    WHERE fcm.food_id = %s
                      AND fcm.source_id = %s
                      AND fcm.category_id = fc.category_id
                      AND fc.source_system = 'ciqual_2025'
                    """,
                    (food_id, source_id),
                )
                cur.execute(
                    """
                    INSERT INTO food_category_memberships (food_id, category_id, source_id, is_primary)
                    VALUES (%s, %s, %s, TRUE)
                    ON CONFLICT (food_id, category_id, source_id)
                    DO UPDATE SET is_primary = EXCLUDED.is_primary
                    """,
                    (food_id, target_category_id, source_id),
                )
                result.categories_assigned += 1
            cur.execute("RELEASE SAVEPOINT sp_category")
        except Exception as exc:
            cur.execute("ROLLBACK TO SAVEPOINT sp_category")
            cur.execute("RELEASE SAVEPOINT sp_category")
            result.errors.append(f"Category assignment {alim_code}: {exc}")

    for const_code, obs in record["nutrients"].items():
        if const_code not in nutrient_ids:
            continue

        if obs["comparator"] == "missing" and obs["amount_raw"] in ("", "-"):
            result.observations_skipped += 1
            continue

        cur.execute("SAVEPOINT sp_observation")
        try:
            source_ref = f"CIQUAL:{alim_code}:{const_code}"
            cur.execute(
                """
                SELECT 1
                FROM food_nutrient_observations
                WHERE source_record_ref = %s AND source_id = %s
                LIMIT 1
                """,
                (source_ref, source_id),
            )
            if cur.fetchone():
                result.observations_skipped += 1
                cur.execute("RELEASE SAVEPOINT sp_observation")
                continue

            cur.execute(
                """
                INSERT INTO food_nutrient_observations
                  (food_id, nutrient_id, source_id, source_record_ref,
                   amount_raw, comparator, amount_value,
                   basis, observed_at, effective_from)
                VALUES (%s, %s, %s, %s, %s, %s, %s, 'per_100g', %s, %s)
                """,
                (
                    food_id,
                    nutrient_ids[const_code],
                    source_id,
                    source_ref,
                    obs["amount_raw"],
                    obs["comparator"],
                    obs["amount_value"],
                    CIQUAL_OBSERVED_AT,
                    CIQUAL_EFFECTIVE_FROM,
                ),
            )
            result.observations_inserted += 1
            cur.execute("RELEASE SAVEPOINT sp_observation")
        except Exception as exc:
            cur.execute("ROLLBACK TO SAVEPOINT sp_observation")
            cur.execute("RELEASE SAVEPOINT sp_observation")
            result.errors.append(f"Observation {alim_code}/{const_code}: {exc}")


def _verify_unresolved_fr(cur, result):
    cur.execute(
        """
        SELECT count(DISTINCT f.food_id)
        FROM foods f
        JOIN food_external_refs fer ON fer.food_id = f.food_id
        WHERE fer.ref_system = 'CIQUAL'
          AND f.canonical_name_fr IS NULL
        """
    )
    result.unresolved_count = cur.fetchone()[0]

    if result.unresolved_count <= 0:
        return

    cur.execute(
        """
        SELECT fer.ref_value, COALESCE(f.canonical_name_en, '')
        FROM foods f
        JOIN food_external_refs fer ON fer.food_id = f.food_id
        WHERE fer.ref_system = 'CIQUAL'
          AND f.canonical_name_fr IS NULL
        ORDER BY fer.ref_value
        LIMIT 20
        """
    )
    samples = cur.fetchall()
    for code, name_en in samples:
        if len(result.unresolved_examples) >= 20:
            break
        result.unresolved_examples.append((str(code), name_en))


def cmd_load(filepath, alim_xml_path, alim_grp_xml_path, db_url):
    import psycopg2

    print("=== CIQUAL 2025 -> Postgres Load ===")
    print(f"XLSX:      {filepath}")
    print(f"ALIM XML:  {alim_xml_path}")
    print(f"GRP XML:   {alim_grp_xml_path}")
    print(f"DB:        {db_url}")
    print()

    wb, ws, metadata = open_ciqual_workbook(filepath)
    try:
        _print_mapping(metadata)
        alim_lookup = load_alim_xml(alim_xml_path)
        grp_nodes = load_alim_grp_xml(alim_grp_xml_path)

        print(f"Loaded XML identity records: {len(alim_lookup)}")
        print(f"Loaded XML category nodes:   {len(grp_nodes)}")
        print()

        conn = psycopg2.connect(db_url)
        conn.autocommit = False
        cur = conn.cursor()
        result = LoadResult()

        try:
            source_id, nutrient_ids = _resolve_source_and_nutrients(cur)
            print(f"Source ID: {source_id}")
            category_ids, result.categories_created, result.categories_updated = seed_ciqual_categories(cur, grp_nodes)

            processed = 0
            for raw_record in iter_ciqual_records(ws, metadata):
                processed += 1
                record = merge_record_identity(raw_record, alim_lookup)
                _load_record(cur, source_id, nutrient_ids, category_ids, record, result)

            conn.commit()
            print(f"Processed food rows:         {processed}")

            _verify_unresolved_fr(cur, result)

            print()
            print("=== Load Summary ===")
            print(f"  Foods created:               {result.foods_created}")
            print(f"  Foods already existing:      {result.foods_existing}")
            print(f"  Foods unresolved FR:         {result.foods_unresolved_fr}")
            print(f"  Categories created:          {result.categories_created}")
            print(f"  Categories updated:          {result.categories_updated}")
            print(f"  Category memberships set:    {result.categories_assigned}")
            print(f"  Group-00 rows (no category): {result.aggregate_rows_skipped}")
            print(f"  Observations inserted:       {result.observations_inserted}")
            print(f"  Observations skipped:        {result.observations_skipped}")
            print(f"  Errors:                      {len(result.errors)}")
            if result.errors:
                for err in result.errors[:20]:
                    print(f"    - {err}")
                if len(result.errors) > 20:
                    print(f"    ... and {len(result.errors) - 20} more")

            print()
            print("=== Post-load verification ===")
            cur.execute(
                """
                SELECT derivation_status, count(*)
                FROM v_food_excess_fructose_latest
                GROUP BY derivation_status
                ORDER BY count(*) DESC
                """
            )
            print("  Excess fructose derivation:")
            for status, cnt in cur.fetchall():
                print(f"    {status}: {cnt}")

            print()
            print(f"  Unresolved CIQUAL foods (canonical_name_fr IS NULL): {result.unresolved_count}")
            if result.unresolved_count > 0:
                print("  Sample unresolved codes:")
                for code, name_en in result.unresolved_examples[:20]:
                    suffix = f" ({name_en})" if name_en else ""
                    print(f"    - {code}{suffix}")
                print("  FAIL: unresolved FR names remain after load.")
                result.exit_code = 2
            else:
                print("  OK: no unresolved FR names in CIQUAL-linked foods.")

            cur.close()
            conn.close()
            return result.exit_code

        except Exception:
            conn.rollback()
            cur.close()
            conn.close()
            raise

    finally:
        wb.close()


def main():
    parser = argparse.ArgumentParser(description="CIQUAL 2025 -> FODMAP Planner ETL")
    parser.add_argument(
        "command",
        choices=["inspect", "stats", "load"],
        help="inspect: file structure; stats: nutrient coverage; load: insert to Postgres",
    )
    parser.add_argument("filepath", type=Path, help="Path to CIQUAL 2025 English XLSX file")
    parser.add_argument("--alim-xml", type=Path, help="Path to alim_*.xml (required for load)")
    parser.add_argument("--alim-grp-xml", type=Path, help="Path to alim_grp_*.xml (required for load)")
    parser.add_argument("--db-url", default=DEFAULT_DB_URL, help=f"Postgres URL (default: {DEFAULT_DB_URL})")
    args = parser.parse_args()

    if not args.filepath.exists():
        print(f"Error: file not found: {args.filepath}")
        sys.exit(1)

    if args.command == "inspect":
        cmd_inspect(args.filepath)
        return

    if args.command == "stats":
        cmd_stats(args.filepath)
        return

    if args.command == "load":
        if args.alim_xml is None or args.alim_grp_xml is None:
            print("Error: load requires --alim-xml and --alim-grp-xml")
            sys.exit(1)
        if not args.alim_xml.exists():
            print(f"Error: file not found: {args.alim_xml}")
            sys.exit(1)
        if not args.alim_grp_xml.exists():
            print(f"Error: file not found: {args.alim_grp_xml}")
            sys.exit(1)
        rc = cmd_load(args.filepath, args.alim_xml, args.alim_grp_xml, args.db_url)
        sys.exit(rc)


if __name__ == "__main__":
    main()
